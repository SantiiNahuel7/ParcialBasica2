from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# Crear libro de Excel
wb = Workbook()

# ------------------ Hoja 1: Parametrización de Médicos ------------------
ws1 = wb.active
ws1.title = "Parametrización de Médicos"
headers1 = ["Médico", "DNI", "Zona", "Días de Guardia", "Horario Inicio", "Horario Fin", "Tipo de Guardia"]
ws1.append(headers1)

# ------------------ Hoja 2: Registro Diario ------------------
ws2 = wb.create_sheet("Registro Diario")
headers2 = ["Fecha", "Médico", "Zona", "Estado", "Hora Inicio", "Hora Fin", "Observaciones"]
ws2.append(headers2)

# Validación de datos para Estado
estado_options = ["Cumplió", "Ausente", "Baja por demanda", "Recorte"]
dv_estado = DataValidation(type="list", formula1=f'"{','.join(estado_options)}"', allow_blank=True)
dv_estado.prompt = "Selecciona un estado"
dv_estado.promptTitle = "Estado"
ws2.add_data_validation(dv_estado)
dv_estado.add(f"D2:D1048576")

# Validación de datos para Zona
zona_options = ["CABA", "Zona Oeste", "Zona Sur", "Zona Norte"]
dv_zona = DataValidation(type="list", formula1=f'"{','.join(zona_options)}"', allow_blank=True)
dv_zona.prompt = "Selecciona una zona"
dv_zona.promptTitle = "Zona"
ws2.add_data_validation(dv_zona)
dv_zona.add(f"C2:C1048576")

# Formato condicional para destacar Ausente y Recorte
fill_ausente = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_recorte = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

rule_ausente = FormulaRule(formula=["$D2=\"Ausente\""], fill=fill_ausente)
rule_recorte = FormulaRule(formula=["$D2=\"Recorte\""], fill=fill_recorte)

ws2.conditional_formatting.add("A2:G1048576", rule_ausente)
ws2.conditional_formatting.add("A2:G1048576", rule_recorte)

# ------------------ Hoja 3: Pronóstico de Demanda ------------------
ws3 = wb.create_sheet("Pronóstico de Demanda")
headers3 = ["Fecha", "Zona", "Demanda Estimada", "Recomendación"]
ws3.append(headers3)

# Validación de datos para Zona en hoja 3
ws3.add_data_validation(dv_zona)
dv_zona.add(f"B2:B1048576")

# ------------------ Hoja 4: Alertas y Notificaciones ------------------
ws4 = wb.create_sheet("Alertas y Notificaciones")
headers4 = ["Fecha", "Médico", "Zona", "Motivo", "Avisado (Sí/No)"]
ws4.append(headers4)

# Validación de datos para Zona en hoja 4
ws4.add_data_validation(dv_zona)
dv_zona.add(f"C2:C1048576")

# Guardar archivo
output_path = "/mnt/data/Plantilla_Grilla_Guardias.xlsx"
wb.save(output_path)
